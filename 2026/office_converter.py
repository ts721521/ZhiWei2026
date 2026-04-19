# -*- coding: utf-8 -*-
"""
Office batch conversion and file organizing tool - core logic

Notes:
- Can run as standalone CLI and can also be invoked by GUI (office_gui.py).
- In GUI mode, use interactive=False to disable CLI input() prompts.
"""

import os
import sys
import time
import json
import csv
import shutil
import logging
import argparse
import uuid
import traceback
import tempfile
import subprocess
import threading
import signal
import random
import hashlib
import re
from datetime import datetime
from pathlib import Path

# Avoid UnicodeEncodeError on Windows consoles with legacy code pages.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(errors="replace")
    except (AttributeError, OSError, RuntimeError, ValueError):
        pass

# win32com is required on Windows; use mocks on non-Windows environments.
HAS_WIN32 = False
try:
    import win32com.client
    import pythoncom
    import pywintypes

    HAS_WIN32 = True
except ImportError:
    # Provide mock objects on non-Windows environments to avoid import failures.
    class MockComError(Exception):
        pass

    class MockPyWinTypes:
        com_error = MockComError

    class MockPythonCom:
        def CoInitialize(self):
            pass

        def CoUninitialize(self):
            pass

    class MockWin32ComClient:
        def Dispatch(self, *args, **kwargs):
            raise RuntimeError("Win32 COM not supported")

        def DispatchEx(self, *args, **kwargs):
            raise RuntimeError("Win32 COM not supported")

    class MockWin32Com:
        client = MockWin32ComClient()

    pywintypes = MockPyWinTypes()
    pythoncom = MockPythonCom()
    win32com = MockWin32Com()


# For timed keyboard input on Windows (retry prompt in CLI mode).
try:
    import msvcrt

    HAS_MSVCRT = True
except ImportError:
    HAS_MSVCRT = False

# pypdf is used for PDF merging and text extraction/scanning.
try:
    from pypdf import PdfWriter, PdfReader

    HAS_PYPDF = True
except ImportError:
    HAS_PYPDF = False

# openpyxl is used to generate Excel indexes in collect_only mode.
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import chromadb

    HAS_CHROMADB = True
except ImportError:
    HAS_CHROMADB = False

try:
    from bs4 import BeautifulSoup

    HAS_BS4 = True
except ImportError:
    HAS_BS4 = False

try:
    from docx import Document

    HAS_PYDOCX = True
except ImportError:
    HAS_PYDOCX = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

try:
    from markitdown import MarkItDown

    HAS_MARKITDOWN = True
except ImportError:
    HAS_MARKITDOWN = False

__version__ = "5.21.2"

from converter.constants import (
    wdFormatPDF,
    xlTypePDF,
    ppSaveAsPDF,
    ppFixedFormatTypePDF,
    xlPDF_SaveAs,
    xlRepairFile,
    ENGINE_WPS,
    ENGINE_MS,
    ENGINE_ASK,
    KILL_MODE_ASK,
    KILL_MODE_AUTO,
    KILL_MODE_KEEP,
    MODE_CONVERT_ONLY,
    MODE_MERGE_ONLY,
    MODE_CONVERT_THEN_MERGE,
    MODE_COLLECT_ONLY,
    MODE_MSHELP_ONLY,
    MERGE_CONVERT_SUBMODE_MERGE_ONLY,
    MERGE_CONVERT_SUBMODE_PDF_TO_MD,
    COLLECT_COPY_LAYOUT_FLAT,
    COLLECT_COPY_LAYOUT_PRESERVE_TREE,
    COLLECT_MODE_COPY_AND_INDEX,
    COLLECT_MODE_INDEX_ONLY,
    MERGE_MODE_CATEGORY,
    MERGE_MODE_ALL_IN_ONE,
    STRATEGY_STANDARD,
    STRATEGY_SMART_TAG,
    STRATEGY_PRICE_ONLY,
    ERR_RPC_SERVER_BUSY,
    ERR_RPC_SERVER_UNAVAILABLE,
)



from converter.errors import ConversionErrorType, classify_conversion_error
from converter.default_config import create_default_config
from converter.config_load import load_config as load_config_impl
from converter.platform_utils import clear_console, get_app_path, is_mac, is_win
from converter.file_registry import FileRegistry
from converter.output_plan import compute_convert_output_plan
from converter.config_defaults import apply_config_defaults_for_converter
from converter.readable import (
    readable_collect_mode,
    readable_content_strategy,
    readable_engine_type,
    readable_merge_mode,
    readable_run_mode,
)
from converter.runtime_prefs import get_merge_convert_submode, get_output_pref
from converter.path_config import (
    get_path_from_config,
    init_paths_from_config_for_converter,
    save_config as save_config_impl,
)
from converter.office_cycle import (
    get_app_type_for_ext,
    get_office_restart_every,
    should_reuse_office_app,
)
from converter.process_policy import resolve_process_handling
from converter.config_terminal import (
    confirm_config_in_terminal as confirm_config_in_terminal_impl,
)
from converter.interactive_choices import (
    ask_for_subfolder as ask_for_subfolder_impl,
    select_collect_mode as select_collect_mode_impl,
    select_content_strategy as select_content_strategy_impl,
    select_engine_mode as select_engine_mode_impl,
    select_merge_mode as select_merge_mode_impl,
    select_run_mode as select_run_mode_impl,
)
from converter.perf_summary import (
    add_perf_seconds as add_perf_seconds_impl,
    build_perf_summary,
)
from converter.process_ops import (
    kill_process_by_name_for_converter,
    process_names_for_engine,
)
from converter.local_office_app import get_local_app as get_local_app_impl
from converter.mac_convert import convert_on_mac as convert_on_mac_impl
from converter.target_path import get_target_path as get_target_path_impl
from converter.pdf_content_scan import scan_pdf_content as scan_pdf_content_impl
from converter.error_summary import (
    get_error_summary_for_display as get_error_summary_for_display_impl,
)
from converter.mshelp_workflow import run_mshelp_only as run_mshelp_only_impl
from converter.runtime_summary import print_runtime_summary as print_runtime_summary_impl
from converter.bootstrap_state import (
    build_default_perf_metrics as build_default_perf_metrics_impl,
    handle_stop_signal as handle_stop_signal_impl,
    initialize_converter_for_runtime,
)
from converter.runtime_lifecycle import (
    check_and_handle_running_processes_for_converter,
    cleanup_all_processes as cleanup_all_processes_impl,
    close_office_apps as close_office_apps_impl,
    kill_current_app as kill_current_app_impl,
    on_office_file_processed as on_office_file_processed_impl,
)
from converter.cli_wizard_flow import run_cli_wizard as run_cli_wizard_impl
from converter.display_helpers import (
    print_step_title as print_step_title_impl,
    print_welcome as print_welcome_impl,
)
from converter.hash_utils import (
    build_short_id,
    compute_file_hash,
    compute_md5,
    make_file_hyperlink,
    mask_md5,
)
from converter.excel_json_utils import (
    build_column_profiles,
    col_index_to_label,
    detect_json_value_type,
    extract_formula_sheet_refs,
    is_effectively_empty_row,
    is_empty_json_cell,
    json_safe_value,
    looks_like_header_row,
    normalize_header_row,
)
from converter.artifact_meta import add_artifact, safe_file_meta
from converter.failure_stage import (
    get_failure_output_expectation,
    infer_failure_stage,
    sanitize_failure_log_stem,
)
from converter.failure_trace_utils import (
    build_failed_file_trace_payload,
    write_failed_file_trace_log,
)
from converter.chromadb_utils import (
    chunk_text_for_vector,
    resolve_chromadb_persist_dir,
    sanitize_chromadb_collection_name,
)
from converter.markdown_text_utils import (
    clean_markdown_page_lines,
    collect_margin_candidates,
    looks_like_heading_line,
    looks_like_page_number_line,
    normalize_extracted_text,
    normalize_margin_line,
    render_markdown_blocks,
)
from converter.naming_utils import ext_bucket, format_merge_filename
from converter.excel_chart_utils import extract_chart_title_text, stringify_chart_anchor
from converter.ai_paths import build_ai_output_path, build_ai_output_path_from_source
from converter.ai_paths_runtime import (
    build_ai_output_path_from_source_for_converter as build_ai_output_path_from_source_for_converter_impl,
)
from converter.text_helpers import (
    extract_mshc_payload,
    find_files_recursive,
    meta_content_by_names,
    normalize_md_line,
    wrap_plain_text_for_pdf,
)
from converter.excel_sheet_utils import auto_fit_sheet, style_header_row
from converter.merge_candidates import (
    build_markdown_merge_tasks,
    scan_merge_candidates_by_ext_for_converter,
)
from converter.merge_tasks import get_merge_tasks as get_merge_tasks_impl
from converter.merge_markdown import merge_markdowns as merge_markdowns_impl
from converter.merge_mode_pipeline import (
    run_merge_mode_pipeline as run_merge_mode_pipeline_impl,
)
from converter.checkpoint_utils import (
    clear_checkpoint_file,
    get_checkpoint_path,
    mark_file_done_in_checkpoint,
    save_checkpoint,
)
from converter.checkpoint_runtime import init_checkpoint as init_checkpoint_impl
from converter.runtime_paths import (
    resolve_incremental_registry_path,
    resolve_update_package_root,
)
from converter.batch_helpers import collect_retry_candidates, get_progress_prefix
from converter.callback_utils import emit_file_done, emit_file_plan
from converter.batch_parallel import run_batch_parallel as run_batch_parallel_impl
from converter.batch_parallel import convert_single_file_threadsafe_for_converter
from converter.batch_sequential import run_batch as run_batch_impl
from converter.process_single import process_single_file as process_single_file_impl
from converter.run_workflow import run as run_workflow_impl
from converter.convert_thread import convert_logic_in_thread as convert_logic_in_thread_impl
from converter.convert_thread_runtime import (
    convert_logic_in_thread_for_converter as convert_logic_in_thread_for_converter_impl,
)
from converter.corpus_manifest import (
    maybe_build_llm_delivery_hub as maybe_build_llm_delivery_hub_impl,
)
from converter.corpus_manifest import write_corpus_manifest as write_corpus_manifest_impl
from converter.collect_index import (
    collect_office_files_and_build_excel as collect_office_files_and_build_excel_impl,
)
from converter.merge_pdfs import merge_pdfs as merge_pdfs_impl
from converter.source_roots import (
    get_configured_source_roots,
    get_source_root_for_path,
    get_source_roots,
    probe_source_root_access,
)
from converter.safe_exec import safe_exec
from converter.file_ops import (
    copy_pdf_direct,
    handle_file_conflict,
    quarantine_failed_file,
    unblock_file,
)
from converter.interactive_prompts import confirm_continue_missing_md_merge
from converter.incremental_filters import (
    apply_global_md5_dedup,
    apply_source_priority_filter,
)
from converter.incremental_registry_ops import (
    build_source_meta,
    flush_incremental_registry,
)
from converter.incremental_scan import apply_incremental_filter_for_converter
from converter.scan_convert_candidates import scan_convert_candidates
from converter.mshelp_scan import find_mshelpviewer_dirs, scan_mshelp_cab_candidates
from converter.mshelp_topics import parse_mshelp_topics
from converter.update_package_index import write_update_package_index_xlsx
from converter.mshelp_records import (
    append_mshelp_record,
    write_mshelp_index_files,
)
from converter.cab_extract import extract_cab_with_fallback as extract_cab_with_fallback_impl
from converter.cab_convert import convert_cab_to_markdown as convert_cab_to_markdown_impl
from converter.merge_index_doc import (
    create_index_doc_and_convert as create_index_doc_and_convert_impl,
)
from converter.index_runtime import (
    write_merge_map_for_converter,
    write_conversion_index_workbook_for_converter,
    append_conversion_index_record as append_conversion_index_record_impl,
)
from converter.failure_report import (
    export_failed_files_report_for_converter,
)
from converter.excel_page_setup import setup_excel_pages as setup_excel_pages_impl
from converter.excel_chart_extract import (
    extract_sheet_charts as extract_sheet_charts_impl,
    extract_sheet_pivot_tables as extract_sheet_pivot_tables_impl,
)
from converter.excel_content_scan import (
    scan_excel_content_in_thread as scan_excel_content_in_thread_impl,
)
from converter.excel_json_batch import (
    write_excel_structured_json_exports as write_excel_structured_json_exports_impl,
)
from converter.error_recording import (
    record_detailed_error as record_detailed_error_impl,
    record_scan_access_skip as record_scan_access_skip_impl,
)
from converter.traceability import (
    write_trace_map_for_converter as write_trace_map_for_converter_impl,
)
from converter.fast_md_engine import run_fast_md_pipeline as run_fast_md_pipeline_impl
from converter.prompt_wrapper import (
    write_prompt_ready_for_converter,
)
from converter.markdown_source_reader import (
    convert_source_to_markdown_text as convert_source_to_markdown_text_impl,
)
from converter.chromadb_docs import collect_chromadb_documents
from converter.markdown_docx_export import (
    export_markdown_to_docx as export_markdown_to_docx_impl,
)
from converter.markdown_pdf_export import (
    export_markdown_to_pdf as export_markdown_to_pdf_impl,
)
from converter.markdown_render import (
    render_html_to_markdown as render_html_to_markdown_impl,
    table_to_markdown_lines as table_to_markdown_lines_impl,
)
from converter.markdown_quality_report import (
    write_markdown_quality_report_for_converter,
)
from converter.pdf_markdown_export import (
    export_pdf_markdown as export_pdf_markdown_impl,
)
from converter.pdf_markdown_runtime import (
    export_pdf_markdown_for_converter as export_pdf_markdown_for_converter_impl,
)
from converter.chromadb_export import write_chromadb_export as write_chromadb_export_impl
from converter.chromadb_runtime import (
    write_chromadb_export_for_converter as write_chromadb_export_for_converter_impl,
)
from converter.mshelp_merge import (
    merge_mshelp_markdowns_for_converter as merge_mshelp_markdowns_impl_for_converter,
)
from converter.excel_json_export import (
    export_single_excel_json as export_single_excel_json_impl,
)
from converter.excel_defined_names import (
    extract_workbook_defined_names as extract_workbook_defined_names_impl,
)
from converter.index_sheets import (
    write_conversion_index_sheet as write_conversion_index_sheet_impl,
    write_merge_index_sheet as write_merge_index_sheet_impl,
)
from converter.sandbox_guard import (
    check_sandbox_free_space_or_raise as check_sandbox_free_space_or_raise_impl,
)
from converter.records_json_export import (
    write_records_json_exports_for_converter,
)
from converter.logging_setup import setup_logging as setup_logging_impl
from converter.retry_prompt import ask_retry_failed_files as ask_retry_failed_files_impl
from converter.update_package_export import (
    generate_update_package_for_converter,
)


class OfficeConverter:
    def __init__(self, config_path: str, interactive: bool = True):
        """Initialize converter."""
        initialize_converter_for_runtime(
            self,
            config_path=config_path,
            interactive=interactive,
            mode_convert_then_merge=MODE_CONVERT_THEN_MERGE,
            collect_mode_copy_and_index=COLLECT_MODE_COPY_AND_INDEX,
            merge_mode_category=MERGE_MODE_CATEGORY,
            strategy_standard=STRATEGY_STANDARD,
            signal_module=signal,
            current_thread_fn=threading.current_thread,
            main_thread_fn=threading.main_thread,
        )

    def _reset_perf_metrics(self):
        self.perf_metrics = build_default_perf_metrics_impl()

    def _add_perf_seconds(self, key, seconds):
        return add_perf_seconds_impl(self.perf_metrics, key, seconds)

    # =============== Base initialization ===============

    def _init_paths_from_config(self):
        return init_paths_from_config_for_converter(
            self,
            get_app_path_fn=get_app_path,
            gettempdir_fn=tempfile.gettempdir,
            isabs_fn=os.path.isabs,
            abspath_fn=os.path.abspath,
            join_fn=os.path.join,
            makedirs_fn=os.makedirs,
        )

    # =============== Shared display helpers ===============

    def print_welcome(self):
        return print_welcome_impl(
            app_version=__version__,
            config_path=self.config_path,
            print_fn=print,
        )

    def print_step_title(self, text):
        return print_step_title_impl(text, print_fn=print)

    def signal_handler(self, signum, _frame):
        return handle_stop_signal_impl(
            signum,
            set_running_fn=lambda flag: setattr(self, "is_running", flag),
            log_warning_fn=logging.warning,
        )

    def get_readable_run_mode(self):
        return readable_run_mode(self.run_mode)

    def get_readable_collect_mode(self):
        return readable_collect_mode(self.collect_mode)

    def get_readable_content_strategy(self):
        return readable_content_strategy(self.content_strategy)

    def get_readable_engine_type(self):
        return readable_engine_type(self.engine_type)

    def get_readable_merge_mode(self):
        return readable_merge_mode(self.merge_mode)

    compute_convert_output_plan = staticmethod(compute_convert_output_plan)

    def _get_output_pref(self):
        return get_output_pref(self.config)

    def _get_merge_convert_submode(self):
        return get_merge_convert_submode(self.config)

    def print_runtime_summary(self):
        return print_runtime_summary_impl(
            config=self.config,
            run_mode=self.run_mode,
            merge_mode=self.merge_mode,
            content_strategy=self.content_strategy,
            mode_merge_only=MODE_MERGE_ONLY,
            get_output_pref_fn=self._get_output_pref,
            get_merge_convert_submode_fn=self._get_merge_convert_submode,
            should_reuse_office_app_fn=self._should_reuse_office_app,
            get_office_restart_every_fn=self._get_office_restart_every,
            print_fn=print,
        )

    def cleanup_all_processes(self):
        return cleanup_all_processes_impl(
            self.engine_type,
            process_names_for_engine_fn=process_names_for_engine,
            kill_process_by_name_fn=self._kill_process_by_name,
        )

    def _kill_process_by_name(self, app_name):
        return kill_process_by_name_for_converter(
            app_name,
            has_win32=HAS_WIN32,
            run_cmd=subprocess.run,
        )

    def _check_sandbox_free_space_or_raise(self):
        return check_sandbox_free_space_or_raise_impl(
            self.config,
            exists_fn=os.path.exists,
            splitdrive_fn=os.path.splitdrive,
            getcwd_fn=os.getcwd,
            disk_usage_fn=shutil.disk_usage,
            log_info_fn=logging.info,
            log_warning_fn=logging.warning,
            print_fn=print,
        )

    def load_config(self, path):
        return load_config_impl(
            self,
            path,
            open_fn=open,
            json_loads_fn=json.loads,
            abspath_fn=os.path.abspath,
            print_fn=print,
            exit_fn=sys.exit,
        )

    def _apply_config_defaults(self):
        return apply_config_defaults_for_converter(self)

    def _should_reuse_office_app(self):
        return should_reuse_office_app(
            self.config, has_win32=HAS_WIN32, is_mac_platform=is_mac()
        )

    def _get_office_restart_every(self):
        return get_office_restart_every(self.config)

    def _get_app_type_for_ext(self, ext):
        return get_app_type_for_ext(self.config, ext)

    def _on_office_file_processed(self, ext):
        self._office_file_counter = on_office_file_processed_impl(
            ext,
            should_reuse_office_app_fn=self._should_reuse_office_app,
            reuse_process=self.reuse_process,
            get_office_restart_every_fn=self._get_office_restart_every,
            get_app_type_for_ext_fn=self._get_app_type_for_ext,
            office_file_counter=self._office_file_counter,
            kill_current_app_fn=self._kill_current_app,
            log_info_fn=logging.info,
        )

    def _get_path_from_config(self, key_base):
        return get_path_from_config(
            self.config,
            key_base,
            prefer_win=is_win(),
            prefer_mac=is_mac(),
        )

    def save_config(self):
        return save_config_impl(
            self.config_path,
            self.config,
            open_fn=open,
            dump_fn=json.dump,
            log_error_fn=logging.error,
        )

    def cli_wizard(self):
        return run_cli_wizard_impl(
            interactive=self.interactive,
            print_welcome_fn=self.print_welcome,
            confirm_config_in_terminal_fn=self.confirm_config_in_terminal,
            ask_for_subfolder_fn=self.ask_for_subfolder,
            select_run_mode_fn=self.select_run_mode,
            get_run_mode_fn=lambda: self.run_mode,
            select_collect_mode_fn=self.select_collect_mode,
            select_content_strategy_fn=self.select_content_strategy,
            select_merge_mode_fn=self.select_merge_mode,
            select_engine_mode_fn=self.select_engine_mode,
            check_and_handle_running_processes_fn=self.check_and_handle_running_processes,
            init_paths_from_config_fn=self._init_paths_from_config,
            config=self.config,
            mode_collect_only=MODE_COLLECT_ONLY,
            mode_convert_only=MODE_CONVERT_ONLY,
            mode_convert_then_merge=MODE_CONVERT_THEN_MERGE,
            mode_merge_only=MODE_MERGE_ONLY,
            mode_mshelp_only=MODE_MSHELP_ONLY,
        )

    def check_and_handle_running_processes(self):
        return check_and_handle_running_processes_for_converter(
            self,
            resolve_process_handling_fn=resolve_process_handling,
        )

    def confirm_config_in_terminal(self):
        return confirm_config_in_terminal_impl(self)

    def ask_for_subfolder(self):
        return ask_for_subfolder_impl(
            self.config,
            print_step_title_fn=self.print_step_title,
            print_fn=print,
            input_fn=input,
        )

    def select_run_mode(self):
        self.run_mode = select_run_mode_impl(
            print_step_title_fn=self.print_step_title,
            get_readable_run_mode_fn=readable_run_mode,
            mode_convert_only=MODE_CONVERT_ONLY,
            mode_merge_only=MODE_MERGE_ONLY,
            mode_convert_then_merge=MODE_CONVERT_THEN_MERGE,
            mode_collect_only=MODE_COLLECT_ONLY,
            mode_mshelp_only=MODE_MSHELP_ONLY,
            print_fn=print,
            input_fn=input,
        )

    def select_collect_mode(self):
        self.collect_mode = select_collect_mode_impl(
            print_step_title_fn=self.print_step_title,
            get_readable_collect_mode_fn=readable_collect_mode,
            collect_mode_copy_and_index=COLLECT_MODE_COPY_AND_INDEX,
            collect_mode_index_only=COLLECT_MODE_INDEX_ONLY,
            print_fn=print,
            input_fn=input,
        )

    def select_merge_mode(self):
        self.merge_mode = select_merge_mode_impl(
            self.config,
            print_step_title_fn=self.print_step_title,
            get_readable_merge_mode_fn=readable_merge_mode,
            merge_mode_category=MERGE_MODE_CATEGORY,
            merge_mode_all_in_one=MERGE_MODE_ALL_IN_ONE,
            print_fn=print,
            input_fn=input,
        )

    def select_content_strategy(self):
        self.content_strategy = select_content_strategy_impl(
            self.price_keywords,
            print_step_title_fn=self.print_step_title,
            get_readable_content_strategy_fn=readable_content_strategy,
            strategy_standard=STRATEGY_STANDARD,
            strategy_smart_tag=STRATEGY_SMART_TAG,
            strategy_price_only=STRATEGY_PRICE_ONLY,
            print_fn=print,
            input_fn=input,
        )

    def select_engine_mode(self):
        self.engine_type = select_engine_mode_impl(
            self.config,
            print_step_title_fn=self.print_step_title,
            get_readable_engine_type_fn=readable_engine_type,
            engine_ask=ENGINE_ASK,
            engine_wps=ENGINE_WPS,
            engine_ms=ENGINE_MS,
            print_fn=print,
            input_fn=input,
        )

    def setup_logging(self):
        self.log_path = setup_logging_impl(
            config=self.config,
            engine_type=self.engine_type,
            run_mode=self.run_mode,
            content_strategy=self.content_strategy,
            merge_mode=self.merge_mode,
            temp_sandbox=self.temp_sandbox,
            merge_output_dir=self.merge_output_dir,
            app_version=__version__,
            get_readable_run_mode_fn=self.get_readable_run_mode,
            get_readable_content_strategy_fn=self.get_readable_content_strategy,
            get_readable_merge_mode_fn=self.get_readable_merge_mode,
            should_reuse_office_app_fn=self._should_reuse_office_app,
            get_office_restart_every_fn=self._get_office_restart_every,
            mode_convert_only=MODE_CONVERT_ONLY,
            mode_convert_then_merge=MODE_CONVERT_THEN_MERGE,
            mode_merge_only=MODE_MERGE_ONLY,
            now_fn=datetime.now,
            get_app_path_fn=get_app_path,
            logging_module=logging,
        )

    def _kill_current_app(self, app_type, force=False):
        return kill_current_app_impl(
            app_type,
            reuse_process=self.reuse_process,
            force=force,
            engine_type=self.engine_type,
            engine_wps=ENGINE_WPS,
            engine_ms=ENGINE_MS,
            kill_process_by_name_fn=self._kill_process_by_name,
        )

    def _get_local_app(self, app_type):
        return get_local_app_impl(
            app_type=app_type,
            engine_type=self.engine_type,
            has_win32=HAS_WIN32,
            engine_wps=ENGINE_WPS,
            engine_ms=ENGINE_MS,
            pythoncom_module=pythoncom,
            win32_client=win32com.client,
        )

    def close_office_apps(self):
        return close_office_apps_impl(
            reuse_process=self.reuse_process,
            run_mode=self.run_mode,
            mode_merge_only=MODE_MERGE_ONLY,
            mode_collect_only=MODE_COLLECT_ONLY,
            cleanup_all_processes_fn=self.cleanup_all_processes,
        )

    # =============== Path and conflict handling ===============

    def get_target_path(self, source_file_path, ext, prefix_override=None):
        return get_target_path_impl(
            self.config,
            source_file_path,
            ext,
            prefix_override=prefix_override,
        )

    def handle_file_conflict(self, temp_pdf_path, target_pdf_path):
        return handle_file_conflict(temp_pdf_path, target_pdf_path)

    # =============== Content scanning ===============

    def scan_pdf_content(self, pdf_path):
        return scan_pdf_content_impl(
            pdf_path,
            price_keywords=self.price_keywords,
            has_pypdf=HAS_PYPDF,
            pdf_reader_cls=PdfReader if HAS_PYPDF else None,
            max_pages=5,
        )

    def scan_excel_content_in_thread(self, workbook):
        return scan_excel_content_in_thread_impl(
            workbook,
            price_keywords=self.price_keywords,
            log_info_fn=logging.info,
            log_warning_fn=logging.warning,
        )

    # =============== COM safe execution helpers ===============

    def _safe_exec(self, func, *args, retries=3, **kwargs):
        return safe_exec(
            func,
            *args,
            retries=retries,
            is_running_getter=lambda: bool(self.is_running),
            sleep_fn=time.sleep,
            randint_fn=random.randint,
            com_error_cls=pywintypes.com_error,
            rpc_server_busy_code=ERR_RPC_SERVER_BUSY,
            rpc_retry_codes=(ERR_RPC_SERVER_BUSY, ERR_RPC_SERVER_UNAVAILABLE),
            **kwargs,
        )

    def _unblock_file(self, file_path):
        unblock_file(file_path)

    def _setup_excel_pages(self, workbook):
        setup_excel_pages_impl(workbook)

    # =============== MacOS Automation Support (Stub/Future) ===============

    def _convert_on_mac(self, file_source, sandbox_target_pdf, ext):
        return convert_on_mac_impl(
            file_source,
            sandbox_target_pdf,
            ext,
            is_mac_fn=is_mac,
            log_error_fn=logging.error,
            log_warning_fn=logging.warning,
        )

    # =============== Core conversion ===============

    def convert_logic_in_thread(
        self, file_source, sandbox_target_pdf, ext, result_context
    ):
        return convert_logic_in_thread_for_converter_impl(
            self,
            file_source,
            sandbox_target_pdf,
            ext,
            result_context,
            convert_logic_in_thread_fn=convert_logic_in_thread_impl,
            is_mac_fn=is_mac,
            has_win32=HAS_WIN32,
            engine_wps=ENGINE_WPS,
            wd_format_pdf=wdFormatPDF,
            xl_type_pdf=xlTypePDF,
            pp_save_as_pdf=ppSaveAsPDF,
            pp_fixed_format_type_pdf=ppFixedFormatTypePDF,
            xl_pdf_save_as=xlPDF_SaveAs,
            xl_repair_file=xlRepairFile,
            strategy_standard=STRATEGY_STANDARD,
            strategy_price_only=STRATEGY_PRICE_ONLY,
            pythoncom_module=pythoncom,
            os_module=os,
        )

    def copy_pdf_direct(self, source, temp_target):
        copy_pdf_direct(source, temp_target)

    def quarantine_failed_file(self, source_path, should_copy=True):
        return quarantine_failed_file(source_path, self.failed_dir, should_copy=should_copy)

    _sanitize_failure_log_stem = staticmethod(sanitize_failure_log_stem)

    def _get_failure_output_expectation(self):
        return get_failure_output_expectation(
            self.run_mode,
            self.config,
            self.compute_convert_output_plan,
        )

    def _infer_failure_stage(self, source_path, raw_error="", context=None):
        return infer_failure_stage(
            source_path,
            raw_error=raw_error,
            context=context,
            cab_extensions=self.config.get("allowed_extensions", {}).get("cab", []),
            expected_outputs_getter=self._get_failure_output_expectation,
        )

    def _build_failed_file_trace_payload(
        self,
        *,
        source_path,
        error_detail,
        status,
        elapsed,
        is_retry,
        failed_copy_path=None,
        extra_context=None,
    ):
        return build_failed_file_trace_payload(
            source_path=source_path,
            error_detail=error_detail,
            status=status,
            elapsed=elapsed,
            is_retry=is_retry,
            failed_copy_path=failed_copy_path,
            extra_context=extra_context,
            get_failure_output_expectation_fn=self._get_failure_output_expectation,
            get_readable_run_mode_fn=self.get_readable_run_mode,
            get_readable_engine_type_fn=self.get_readable_engine_type,
            infer_failure_stage_fn=self._infer_failure_stage,
        )

    def _write_failed_file_trace_log(self, payload, failed_copy_path=None):
        return write_failed_file_trace_log(
            payload,
            failed_copy_path=failed_copy_path,
            enable_trace_log=self.config.get("enable_failed_file_trace_log", True),
            failed_dir=self.failed_dir,
            target_folder=self.config.get("target_folder", ""),
            sanitize_stem_fn=self._sanitize_failure_log_stem,
            log_error=logging.error,
        )

    def record_detailed_error(self, source_path, exception, context=None):
        return record_detailed_error_impl(
            source_path,
            exception,
            context=context,
            classify_conversion_error_fn=classify_conversion_error,
            abspath_fn=os.path.abspath,
            basename_fn=os.path.basename,
            now_fn=datetime.now,
            infer_failure_stage_fn=self._infer_failure_stage,
            get_failure_output_expectation_fn=self._get_failure_output_expectation,
            detailed_error_records=self.detailed_error_records,
            stats=self.stats,
        )

    def export_failed_files_report(self, output_dir=None):
        return export_failed_files_report_for_converter(
            self,
            output_dir=output_dir,
            now_fn=datetime.now,
            log_error=logging.error,
        )

    def get_error_summary_for_display(self):
        return get_error_summary_for_display_impl(self.detailed_error_records)

    _find_files_recursive = staticmethod(find_files_recursive)

    def _extract_cab_with_fallback(self, cab_path, extract_dir):
        return extract_cab_with_fallback_impl(
            cab_path,
            extract_dir,
            is_win_fn=is_win,
            run_cmd=subprocess.run,
            find_files_recursive_fn=self._find_files_recursive,
            cab_7z_path=self.config.get("cab_7z_path", ""),
            get_app_path_fn=get_app_path,
            which_fn=shutil.which,
        )

    _extract_mshc_payload = staticmethod(extract_mshc_payload)
    _meta_content_by_names = staticmethod(meta_content_by_names)

    def _parse_mshelp_topics(self, html_root):
        return parse_mshelp_topics(
            html_root,
            find_files_recursive_fn=self._find_files_recursive,
            has_bs4=HAS_BS4,
            beautifulsoup_cls=BeautifulSoup if HAS_BS4 else None,
            meta_content_by_names_fn=self._meta_content_by_names,
        )

    _normalize_md_line = staticmethod(normalize_md_line)

    def _table_to_markdown_lines(self, table_tag):
        return table_to_markdown_lines_impl(
            table_tag,
            normalize_md_line_fn=self._normalize_md_line,
        )

    def _render_html_to_markdown(self, html_path):
        return render_html_to_markdown_impl(
            html_path,
            has_bs4=HAS_BS4,
            beautifulsoup_cls=BeautifulSoup if HAS_BS4 else None,
            normalize_md_line_fn=self._normalize_md_line,
            table_to_markdown_lines_fn=self._table_to_markdown_lines,
        )

    def _convert_cab_to_markdown(self, cab_path, source_path_for_output):
        return convert_cab_to_markdown_impl(
            cab_path,
            source_path_for_output,
            has_bs4=HAS_BS4,
            temp_sandbox=self.temp_sandbox,
            uuid4_hex_fn=lambda: uuid.uuid4().hex,
            extract_cab_with_fallback_fn=self._extract_cab_with_fallback,
            find_files_recursive_fn=self._find_files_recursive,
            extract_mshc_payload_fn=self._extract_mshc_payload,
            parse_mshelp_topics_fn=self._parse_mshelp_topics,
            build_ai_output_path_from_source_fn=self._build_ai_output_path_from_source,
            normalize_md_line_fn=self._normalize_md_line,
            render_html_to_markdown_fn=self._render_html_to_markdown,
            append_mshelp_record_fn=self._append_mshelp_record,
            now_fn=datetime.now,
            generated_markdown_outputs=self.generated_markdown_outputs,
            log_warning_fn=logging.warning,
        )

    def _append_mshelp_record(self, source_cab_path, markdown_path, topic_count):
        return append_mshelp_record(
            self.mshelp_records,
            source_cab_path,
            markdown_path,
            topic_count,
            folder_name=self.config.get("mshelpviewer_folder_name", "MSHelpViewer"),
            get_source_root_for_path_fn=self._get_source_root_for_path,
        )

    def _find_mshelpviewer_dirs(self, root_dir):
        return find_mshelpviewer_dirs(
            root_dir,
            folder_name=self.config.get("mshelpviewer_folder_name", "MSHelpViewer"),
            is_win_fn=is_win,
        )

    def _scan_mshelp_cab_candidates(self):
        return scan_mshelp_cab_candidates(
            self.config,
            self._get_source_roots(),
            find_mshelpviewer_dirs_fn=self._find_mshelpviewer_dirs,
            find_files_recursive_fn=self._find_files_recursive,
            is_win_fn=is_win,
        )

    def _write_mshelp_index_files(self):
        return write_mshelp_index_files(
            self.mshelp_records,
            self.config.get("target_folder", ""),
            generated_mshelp_outputs=self.generated_mshelp_outputs,
            log_info=logging.info,
        )

    _wrap_plain_text_for_pdf = staticmethod(wrap_plain_text_for_pdf)

    def _export_markdown_to_docx(self, md_path, out_docx):
        return export_markdown_to_docx_impl(
            md_path,
            out_docx,
            has_pydocx=HAS_PYDOCX,
            document_cls=Document if HAS_PYDOCX else None,
            re_module=re,
        )

    def _export_markdown_to_pdf(self, md_path, out_pdf):
        return export_markdown_to_pdf_impl(
            md_path,
            out_pdf,
            has_reportlab=HAS_REPORTLAB,
            canvas_cls=canvas.Canvas if HAS_REPORTLAB else None,
            page_size=A4 if HAS_REPORTLAB else None,
            wrap_plain_text_for_pdf_fn=self._wrap_plain_text_for_pdf,
        )

    def _merge_mshelp_markdowns(self):
        return merge_mshelp_markdowns_impl_for_converter(
            self,
            now_fn=datetime.now,
            perf_counter_fn=time.perf_counter,
            log_info=logging.info,
            log_warning=logging.warning,
        )

    def process_single_file(
        self, file_path, target_path_initial, ext, progress_str, is_retry=False
    ):
        return process_single_file_impl(
            self,
            file_path,
            target_path_initial,
            ext,
            progress_str,
            is_retry=is_retry,
        )

    # =============== Batch processing / retry ===============

    def get_progress_prefix(self, current, total):
        return get_progress_prefix(current, total)

    def _emit_file_plan(self, file_list):
        emit_file_plan(
            getattr(self, "file_plan_callback", None),
            file_list,
            warn_func=logging.warning,
        )

    def _emit_file_done(self, record):
        emit_file_done(
            getattr(self, "file_done_callback", None),
            record,
            warn_func=logging.warning,
        )

    def run_batch(self, file_list, is_retry=False, source_alias_map=None):
        return run_batch_impl(
            self,
            file_list,
            is_retry=is_retry,
            source_alias_map=source_alias_map,
        )

    # ========== 鏂偣缁紶鍜屽苟鍙戝鐞嗗姛鑳?==========

    def _get_checkpoint_path(self):
        """Return checkpoint path for current task."""
        return get_checkpoint_path(self.config)

    def _init_checkpoint(self, file_list):
        """Initialize checkpoint and return (checkpoint_data, pending_files)."""
        return init_checkpoint_impl(
            file_list,
            config=self.config,
            get_checkpoint_path_fn=self._get_checkpoint_path,
            checkpoint_resume_callback=getattr(self, "checkpoint_resume_callback", None),
            save_checkpoint_fn=self._save_checkpoint,
            now_fn=datetime.now,
            exists_fn=os.path.exists,
            remove_fn=os.remove,
            open_fn=open,
            print_fn=print,
            log_warning_fn=logging.warning,
        )

    def _save_checkpoint(self, checkpoint):
        """Persist checkpoint to file."""
        save_checkpoint(checkpoint, self._get_checkpoint_path())

    def _mark_file_done_in_checkpoint(self, checkpoint, file_path):
        """Mark a file as completed in checkpoint data."""
        return mark_file_done_in_checkpoint(checkpoint, file_path)

    def _clear_checkpoint(self):
        """娓呴櫎鏂偣鏂囦欢"""
        clear_checkpoint_file(self._get_checkpoint_path())

    def _convert_single_file_threadsafe(
        self, fpath, target_path_initial, ext, progress_prefix, is_retry=False
    ):
        return convert_single_file_threadsafe_for_converter(
            self,
            fpath,
            target_path_initial,
            ext,
            progress_prefix,
            is_retry=is_retry,
            has_win32=HAS_WIN32,
            co_initialize_fn=pythoncom.CoInitialize,
            co_uninitialize_fn=pythoncom.CoUninitialize,
        )

    def run_batch_parallel(self, file_list, is_retry=False, source_alias_map=None):
        return run_batch_parallel_impl(
            self,
            file_list,
            is_retry=is_retry,
            source_alias_map=source_alias_map,
        )

    def _collect_retry_candidates(self):
        return collect_retry_candidates(
            self.failed_dir,
            self.config.get("allowed_extensions", {}),
            self.error_records,
            self.detailed_error_records,
        )

    def ask_retry_failed_files(self, failed_count, timeout=20):
        return ask_retry_failed_files_impl(
            failed_count,
            error_records=self.error_records,
            timeout=timeout,
            has_msvcrt=HAS_MSVCRT,
            msvcrt_module=msvcrt if HAS_MSVCRT else None,
            time_module=time,
            input_fn=input,
            print_fn=print,
        )

    def _create_index_doc_and_convert(self, word_app, file_list, title):
        return create_index_doc_and_convert_impl(
            word_app,
            file_list,
            title,
            temp_sandbox=self.temp_sandbox,
            uuid4_hex_fn=lambda: uuid.uuid4().hex,
            log_error_fn=logging.error,
        )

    _format_merge_filename = staticmethod(format_merge_filename)

    def _get_merge_tasks(self):
        return get_merge_tasks_impl(
            run_mode=self.run_mode,
            merge_source=self.config.get("merge_source", "source"),
            target_folder=self.config["target_folder"],
            get_source_roots_fn=self._get_source_roots,
            failed_dir=self.failed_dir,
            merge_output_dir=self.merge_output_dir,
            merge_mode=self.merge_mode,
            merge_mode_all_in_one=MERGE_MODE_ALL_IN_ONE,
            mode_merge_only=MODE_MERGE_ONLY,
            merge_filename_pattern=self.config.get("merge_filename_pattern"),
            max_merge_size_mb=self.config.get("max_merge_size_mb", 80),
            now_fn=datetime.now,
            format_merge_filename_fn=self._format_merge_filename,
            print_fn=print,
        )

    _compute_md5 = staticmethod(compute_md5)
    _mask_md5 = staticmethod(mask_md5)
    _build_short_id = staticmethod(build_short_id)

    def _write_merge_map(self, output_path, records):
        return write_merge_map_for_converter(
            self,
            output_path,
            records,
            csv_module=csv,
            json_module=json,
            open_fn=open,
        )

    def merge_pdfs(self):
        return merge_pdfs_impl(
            self,
            has_pypdf=HAS_PYPDF,
            has_openpyxl=HAS_OPENPYXL,
            pdf_writer_cls=PdfWriter if HAS_PYPDF else None,
            pdf_reader_cls=PdfReader if HAS_PYPDF else None,
            workbook_cls=Workbook if HAS_OPENPYXL else None,
            pythoncom_mod=pythoncom,
            win32_client=win32com.client,
        )

    def _scan_merge_candidates_by_ext(self, ext):
        return scan_merge_candidates_by_ext_for_converter(
            self,
            ext,
            mode_merge_only=MODE_MERGE_ONLY,
        )

    def _build_markdown_merge_tasks(self, md_files):
        return build_markdown_merge_tasks(md_files, self.merge_mode)

    def merge_markdowns(self, candidates=None):
        return merge_markdowns_impl(self, candidates=candidates)

    def _confirm_continue_missing_md_merge(self):
        return confirm_continue_missing_md_merge(
            self.interactive,
            input_fn=input,
            warn_func=logging.warning,
        )

    def _run_merge_mode_pipeline(self, batch_results):
        return run_merge_mode_pipeline_impl(self, batch_results)

    # =============== File indexing and dedup ==================

    _compute_file_hash = staticmethod(compute_file_hash)
    _make_file_hyperlink = staticmethod(make_file_hyperlink)

    def _append_conversion_index_record(self, source_path, pdf_path, status=""):
        return append_conversion_index_record_impl(
            source_path,
            pdf_path,
            status=status,
            exists_fn=os.path.exists,
            abspath_fn=os.path.abspath,
            basename_fn=os.path.basename,
            relpath_fn=os.path.relpath,
            compute_md5_fn=self._compute_md5,
            get_source_root_for_path_fn=self._get_source_root_for_path,
            target_folder=self.config["target_folder"],
            conversion_index_records=self.conversion_index_records,
        )

    @staticmethod
    def _style_header_row(ws):
        style_header_row(ws, has_openpyxl=HAS_OPENPYXL, font_cls=Font if HAS_OPENPYXL else None)

    @staticmethod
    def _auto_fit_sheet(ws, max_width=90):
        auto_fit_sheet(ws, max_width=max_width)

    def _write_conversion_index_sheet(self, ws, records):
        return write_conversion_index_sheet_impl(
            ws,
            records,
            style_header_row_fn=self._style_header_row,
            auto_fit_sheet_fn=self._auto_fit_sheet,
            make_file_hyperlink_fn=self._make_file_hyperlink,
        )

    def _write_merge_index_sheet(self, ws, records):
        return write_merge_index_sheet_impl(
            ws,
            records,
            style_header_row_fn=self._style_header_row,
            auto_fit_sheet_fn=self._auto_fit_sheet,
            make_file_hyperlink_fn=self._make_file_hyperlink,
        )

    def _write_conversion_index_workbook(self):
        return write_conversion_index_workbook_for_converter(
            self,
            has_openpyxl=HAS_OPENPYXL,
            workbook_cls=Workbook if HAS_OPENPYXL else None,
            now_fn=datetime.now,
            join_path_fn=os.path.join,
            print_fn=print,
            log_warning_fn=logging.warning,
            log_info_fn=logging.info,
        )

    def _build_ai_output_path(self, source_path, sub_dir, ext):
        return build_ai_output_path(
            source_path,
            sub_dir,
            ext,
            self.config.get("target_folder", ""),
        )

    def _build_ai_output_path_from_source(self, source_path, sub_dir, ext):
        return build_ai_output_path_from_source_for_converter_impl(
            self,
            source_path,
            sub_dir,
            ext,
            build_ai_output_path_from_source_fn=build_ai_output_path_from_source,
        )

    _normalize_extracted_text = staticmethod(normalize_extracted_text)
    _normalize_margin_line = staticmethod(normalize_margin_line)
    _looks_like_page_number_line = staticmethod(looks_like_page_number_line)
    _collect_margin_candidates = staticmethod(collect_margin_candidates)
    _clean_markdown_page_lines = staticmethod(clean_markdown_page_lines)
    _looks_like_heading_line = staticmethod(looks_like_heading_line)
    _render_markdown_blocks = staticmethod(render_markdown_blocks)

    _json_safe_value = staticmethod(json_safe_value)
    _is_empty_json_cell = staticmethod(is_empty_json_cell)
    _is_effectively_empty_row = staticmethod(is_effectively_empty_row)
    _looks_like_header_row = staticmethod(looks_like_header_row)
    _normalize_header_row = staticmethod(normalize_header_row)
    _detect_json_value_type = staticmethod(detect_json_value_type)
    _build_column_profiles = staticmethod(build_column_profiles)
    _col_index_to_label = staticmethod(col_index_to_label)
    _extract_formula_sheet_refs = staticmethod(extract_formula_sheet_refs)

    _extract_chart_title_text = staticmethod(extract_chart_title_text)
    _stringify_chart_anchor = staticmethod(stringify_chart_anchor)

    def _extract_sheet_charts(self, ws_formula, series_ref_limit=50):
        return extract_sheet_charts_impl(
            ws_formula,
            extract_chart_title_text_fn=self._extract_chart_title_text,
            stringify_chart_anchor_fn=self._stringify_chart_anchor,
            series_ref_limit=series_ref_limit,
        )

    @staticmethod
    def _extract_sheet_pivot_tables(ws_formula):
        return extract_sheet_pivot_tables_impl(ws_formula)

    @staticmethod
    def _extract_workbook_defined_names(wb_formula):
        return extract_workbook_defined_names_impl(wb_formula)

    def _export_pdf_markdown(self, pdf_path, source_path_hint=None):
        return export_pdf_markdown_for_converter_impl(
            self,
            pdf_path,
            source_path_hint=source_path_hint,
            export_pdf_markdown_fn=export_pdf_markdown_impl,
            has_pypdf=HAS_PYPDF,
            pdf_reader_cls=PdfReader if HAS_PYPDF else None,
            now_fn=datetime.now,
            log_info_fn=logging.info,
            log_error_fn=logging.error,
        )

    def _write_markdown_quality_report(self):
        return write_markdown_quality_report_for_converter(
            self,
            now_fn=datetime.now,
            log_info_fn=logging.info,
        )

    def _write_records_json_exports(self):
        return write_records_json_exports_for_converter(
            self,
            now_fn=datetime.now,
            log_info=logging.info,
        )

    def _write_trace_map(self):
        return write_trace_map_for_converter_impl(
            self,
            has_openpyxl=HAS_OPENPYXL,
            workbook_cls=Workbook if HAS_OPENPYXL else None,
            load_workbook_fn=load_workbook if HAS_OPENPYXL else None,
            font_cls=Font if HAS_OPENPYXL else None,
            style_header_row_fn=self._style_header_row,
            auto_fit_sheet_fn=self._auto_fit_sheet,
            log_warning_fn=logging.warning,
            log_info_fn=logging.info,
        )

    def _convert_source_to_markdown_text(self, source_path):
        return convert_source_to_markdown_text_impl(
            source_path,
            has_markitdown=HAS_MARKITDOWN,
            markitdown_cls=MarkItDown if HAS_MARKITDOWN else None,
            open_fn=open,
        )

    def _run_fast_md_pipeline(self, files, checkpoint=None):
        on_file_done_for_checkpoint = None
        if checkpoint is not None:
            interval = self.config.get("parallel_checkpoint_interval", 10)
            completed_count = [0]

            def _on_done(path):
                self._mark_file_done_in_checkpoint(checkpoint, path)
                completed_count[0] += 1
                if completed_count[0] % interval == 0:
                    self._save_checkpoint(checkpoint)

            on_file_done_for_checkpoint = _on_done

        return run_fast_md_pipeline_impl(
            files,
            config=self.config,
            target_folder=self.config.get("target_folder", ""),
            generated_markdown_outputs=self.generated_markdown_outputs,
            markdown_quality_records=self.markdown_quality_records,
            trace_short_id_taken=self.trace_short_id_taken,
            stats=self.stats,
            source_root_resolver_fn=self._get_source_root_for_path,
            compute_md5_fn=self._compute_md5,
            build_short_id_fn=self._build_short_id,
            convert_source_to_markdown_text_fn=self._convert_source_to_markdown_text,
            progress_callback=getattr(self, "progress_callback", None),
            emit_file_done_fn=self._emit_file_done,
            on_file_done_for_checkpoint=on_file_done_for_checkpoint,
            now_fn=datetime.now,
            log_info_fn=logging.info,
            log_warning_fn=logging.warning,
        )

    def _write_prompt_ready(self):
        return write_prompt_ready_for_converter(
            self,
            now_fn=datetime.now,
            log_info_fn=logging.info,
        )

    _sanitize_chromadb_collection_name = staticmethod(
        sanitize_chromadb_collection_name
    )

    def _resolve_chromadb_persist_dir(self):
        return resolve_chromadb_persist_dir(self.config)

    _chunk_text_for_vector = staticmethod(chunk_text_for_vector)

    def _collect_chromadb_documents(self):
        return collect_chromadb_documents(
            generated_markdown_outputs=self.generated_markdown_outputs,
            markdown_quality_records=self.markdown_quality_records,
            config=self.config,
            chunk_text_for_vector_fn=self._chunk_text_for_vector,
        )

    def _write_chromadb_export(self):
        return write_chromadb_export_for_converter_impl(
            self,
            write_chromadb_export_fn=write_chromadb_export_impl,
            has_chromadb=HAS_CHROMADB,
            chromadb_module=chromadb if HAS_CHROMADB else None,
            now_fn=datetime.now,
            log_info_fn=logging.info,
        )

    def _export_single_excel_json(self, source_excel_path):
        return export_single_excel_json_impl(
            source_excel_path,
            config=self.config,
            build_ai_output_path_from_source_fn=self._build_ai_output_path_from_source,
            source_root_resolver=self._get_source_root_for_path,
            has_openpyxl=HAS_OPENPYXL,
            load_workbook_fn=load_workbook if HAS_OPENPYXL else None,
            extract_workbook_defined_names_fn=self._extract_workbook_defined_names,
            extract_sheet_charts_fn=self._extract_sheet_charts,
            extract_sheet_pivot_tables_fn=self._extract_sheet_pivot_tables,
        )

    def _write_excel_structured_json_exports(self):
        return write_excel_structured_json_exports_impl(
            self,
            log_info_fn=logging.info,
            log_error_fn=logging.error,
        )

    def _safe_file_meta(self, path):
        return safe_file_meta(path, self.config.get("target_folder", ""))

    def _add_artifact(self, artifacts, kind, path):
        add_artifact(artifacts, kind, path, self.config.get("target_folder", ""))

    def _maybe_build_llm_delivery_hub(self, target_folder, artifacts):
        return maybe_build_llm_delivery_hub_impl(self, target_folder, artifacts)

    def _write_corpus_manifest(self, merge_outputs=None):
        return write_corpus_manifest_impl(self, merge_outputs=merge_outputs)

    def collect_office_files_and_build_excel(self):
        return collect_office_files_and_build_excel_impl(
            self,
            has_openpyxl=HAS_OPENPYXL,
            workbook_cls=Workbook if HAS_OPENPYXL else None,
            font_cls=Font if HAS_OPENPYXL else None,
        )

    # =============== Main flow ===============

    def _get_configured_source_roots(self):
        return get_configured_source_roots(self.config)

    def _get_source_roots(self):
        """Return list of accessible source folder paths to scan."""
        return get_source_roots(self.config, is_dir_func=os.path.isdir)

    def _record_scan_access_skip(
        self, path, exception, context=None, seen_keys=None, silent=False
    ):
        return record_scan_access_skip_impl(
            path,
            exception,
            context=context,
            seen_keys=seen_keys,
            silent=silent,
            is_win_fn=is_win,
            abspath_fn=os.path.abspath,
            record_detailed_error_fn=self.record_detailed_error,
            stats=self.stats,
            build_failed_file_trace_payload_fn=self._build_failed_file_trace_payload,
            write_failed_file_trace_log_fn=self._write_failed_file_trace_log,
            print_fn=print,
            log_warning_fn=logging.warning,
        )

    def _probe_source_root_access(self, source_root, context=None, seen_keys=None):
        return probe_source_root_access(
            source_root,
            self._record_scan_access_skip,
            context=context,
            seen_keys=seen_keys,
            is_dir_func=os.path.isdir,
            listdir_fn=os.listdir,
        )

    def _get_source_root_for_path(self, abs_path):
        """Return the source root that contains abs_path (for relpath). Prefer longest match."""
        return get_source_root_for_path(
            abs_path,
            self._get_source_roots(),
            fallback=self.config.get("source_folder", "") or "",
        )

    def _scan_convert_candidates(self):
        return scan_convert_candidates(
            self.config,
            self._get_configured_source_roots(),
            probe_source_root_access_fn=self._probe_source_root_access,
            record_scan_access_skip_fn=self._record_scan_access_skip,
            filter_date=self.filter_date,
            filter_mode=self.filter_mode,
            print_warn_fn=print,
            log_error_fn=logging.error,
        )

    def _apply_source_priority_filter(self, files):
        return apply_source_priority_filter(
            files,
            self.config,
            is_win_fn=is_win,
            log_info=logging.info,
        )

    def _ext_bucket(self, path):
        return ext_bucket(path, self.config.get("allowed_extensions", {}))

    def _apply_global_md5_dedup(self, files):
        return apply_global_md5_dedup(
            files,
            self.config.get("global_md5_dedup", False),
            self._ext_bucket,
            self._compute_md5,
            log_warning=logging.warning,
            log_info=logging.info,
        )

    def _resolve_incremental_registry_path(self):
        return resolve_incremental_registry_path(self.config)

    def _build_source_meta(self, path, include_hash=False):
        return build_source_meta(
            path,
            include_hash=include_hash,
            compute_file_hash_fn=self._compute_file_hash,
            log_warning=logging.warning,
        )

    @staticmethod
    def _incremental_log_info(msg, *args):
        logging.info(msg, *args)

    def _apply_incremental_filter(self, files):
        return apply_incremental_filter_for_converter(self, files)

    def _flush_incremental_registry(self, process_results):
        return flush_incremental_registry(
            self._incremental_context,
            process_results,
            run_mode=self.run_mode,
            compute_md5_fn=self._compute_md5,
            log_info=logging.info,
        )

    def _resolve_update_package_root(self):
        return resolve_update_package_root(self.config)

    def _write_update_package_index_xlsx(self, xlsx_path, records):
        return write_update_package_index_xlsx(
            xlsx_path,
            records,
            has_openpyxl=HAS_OPENPYXL,
            workbook_cls=Workbook if HAS_OPENPYXL else None,
            font_cls=Font if HAS_OPENPYXL else None,
            make_file_hyperlink_fn=self._make_file_hyperlink,
            log_error=logging.error,
        )

    @staticmethod
    def _update_package_log_info(msg):
        logging.info(msg)

    def _generate_update_package(self, process_results):
        return generate_update_package_for_converter(self, process_results)

    def _build_perf_summary(self):
        return build_perf_summary(self.perf_metrics, self.stats)

    def _run_mshelp_only(self):
        return run_mshelp_only_impl(
            stats=self.stats,
            scan_mshelp_cab_candidates_fn=self._scan_mshelp_cab_candidates,
            run_batch_fn=self.run_batch,
            write_mshelp_index_files_fn=self._write_mshelp_index_files,
            merge_mshelp_markdowns_fn=self._merge_mshelp_markdowns,
            add_perf_seconds_fn=self._add_perf_seconds,
            perf_counter_fn=time.perf_counter,
            log_info_fn=logging.info,
            print_fn=print,
        )

    def run(self, resume_file_list=None):
        return run_workflow_impl(self, resume_file_list=resume_file_list, app_version=__version__)


# =============== CLI entry ===============

if __name__ == "__main__":
    clear_console()
    script_dir = get_app_path()
    default_config_path = os.path.join(script_dir, "config.json")
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", default=default_config_path)
    args = parser.parse_args()

    if not os.path.exists(args.config):
        try:
            create_default_config(args.config)
        except (OSError, RuntimeError, TypeError, ValueError):
            pass

    converter = OfficeConverter(args.config, interactive=True)
    converter.cli_wizard()
    converter.run()


